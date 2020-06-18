import numpy as np
import cv2
import os
import glob
from mtcnn.mtcnn import MTCNN
detector = MTCNN()
import openpyxl
#To create an workbook
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title='intensity values'
sheet.cell(row=1, column=1).value='Image_number' #Name of first column
sheet.cell(row=1, column=2).value='Intensity_face'  #Name of second column
sheet.cell(row=1, column=3).value='Intensity_left_cheek'  #Name of third column
sheet.cell(row=1, column=4).value='Intensity_right_cheek'  #Name of fourth column
sheet.cell(row=1, column=5).value='Intensity_forehead'  #Name of fifth column
sheet.cell(row=1, column=6).value='Intensity_nose'  #Name of sixth column


def output(result, thermal_image):
    d=dict() #Dictonary to return the means of ROI's
    bounding_box = result[0]['box']
    keypoints = result[0]['keypoints']
    #print(bounding_box)
    # Points along with box
    xb = bounding_box[0]
    yb = bounding_box[1]-30
    wb = bounding_box[2]
    hb = bounding_box[3]
    # Coordinates of left eye
    xle = keypoints['left_eye'][0]
    yle = keypoints['left_eye'][1]-30
    # Coordinates of right eye
    xre = keypoints['right_eye'][0]
    yre = keypoints['right_eye'][1]-30
    # Coordinates of midpoint of eye
    xe = int(round((xle + xre) / 2, 0))
    ye = int(round((yle + yre) / 2, 0))
    # Coordinates of left mouth
    xlm = keypoints['mouth_left'][0]
    ylm = keypoints['mouth_left'][1]-30
    # Coordinates of right mouth
    xrm = keypoints['mouth_right'][0]
    yrm = keypoints['mouth_right'][1]-30
    #Coordinates of nose
    xn=keypoints['nose'][0]
    yn=keypoints['nose'][1]-30
    # Coordinates of center of mouth
    xm = int(round((xlm + xrm) / 2, 0))
    ym = int(round((ylm + yrm) / 2, 0))
    # Model 1- To identify the cheeks(Yang et al. 2012)
    # Mid point of eye and mouth
    xme = int(round((xe + xm) / 2, 0))
    yme = int(round((ye + ym) / 2, 0))
    # Length of mouth
    Lm = int(xrm - xlm)
    # y coordinates of left and right cheek
    yrc = ylc = yme
    # x Coordinates of left and right cheek
    xlc = int(round((xlm - (Lm) / 3), 0))
    xrc = int(round((xrm + (Lm) / 3), 0))
    # Model-2
    # Distance between center of eyes
    Le = xre - xle
    # Vertical distance between mouth and eye
    Lme = int(ym - ye)
    # Left cheek y coordinate
    yrc2 = ylc2 = int(round(yle + Lme / 4, 0))

    # cv2.circle(image, (263,290), 2, (0, 155, 255), 2)
    # cv2.circle(image, (416,290), 2, (0, 155, 255), 2)
    # Using lines above mouth and below line
    # Point above mouth
    ylam = int(round(ylm - Lme / 4, 0))
    yram = int(round(yrm - Lme / 4, 0))
    pts_left_cheek = np.array([[xle, yrc2], [xlm, ylam], [xb, ylam], [xb, yrc2]], np.int32)
    pts_left_cheek = pts_left_cheek.reshape((-1, 1, 2))

    pts_right_cheek = np.array([[xre, ylc2], [xrm, yram], [xb + wb, yram], [xb + wb, ylc2]], np.int32)
    pts_right_cheek = pts_right_cheek.reshape((-1, 1, 2))
    ylf = int(round(yle - Lme / 3, 0))
    yrf = int(round(yre - Lme / 3, 0))
    pts_forhead = np.array([[xle, ylf], [xle, yb], [xre, yb], [xre, yrf]], np.int32)
    pts_forhead = pts_forhead.reshape(-1, 1, 2)
    crop_img_face = thermal_image[yb:yb+ hb,
                    xb:xb+ wb]
    d['face']=crop_img_face.mean()
    # cv2.rectangle(thermal_image,
    #               (xb, yb),
    #               (xb + w, yb + h),
    #               (0, 155, 255),
    #               2)
    # cv2.circle(thermal_image, (xle,yle), 2, (0, 155, 255), 2)
    # cv2.circle(thermal_image, (xre,yre), 2, (0, 155, 255), 2)
    # cv2.circle(thermal_image, (xn,yn), 2, (0, 155, 255), 2)
    # cv2.circle(thermal_image, (xlm,ylm), 2, (0, 155, 255), 2)
    # cv2.circle(thermal_image, (xrm,yrm), 2, (0, 155, 255), 2)
    wn=int(round(Le/2,0))
    hn=int(round(Lme/2,0))
    xnb=int(round(xn-wn/2,0))
    ynb=int(round(yn-hn,0))
    cv2.rectangle(thermal_image, (xnb,ynb),(xnb+wn,ynb+hn),(0,155,255),2 )
    crop_img_nose = thermal_image[yn:yn + hn, xn:xn + wn].copy()
    d['nose_mean']=crop_img_nose.mean()
    #print(pts_left_cheek)
    if xle - xb < Le / 2: #Left cheek
        rect = cv2.boundingRect(pts_left_cheek)
        x, y, w, h = rect
        crop_img_left_cheek = thermal_image[y:y + h, x:x + w].copy() #Need to make more changes
    else:
        ylc3 = int(round(xle - Le / 2, 0))
        pts_left_cheek = np.array([[xle, yrc2], [xlm, ylam], [ylc3, ylam], [ylc3, yrc2]], np.int32)
        pts_left_cheek = pts_left_cheek.reshape((-1, 1, 2))
        rect = cv2.boundingRect(pts_left_cheek)
        x, y, w, h = rect
        crop_img_left_cheek = thermal_image[y:y + h, x:x + w].copy()  # Need to make more changes.
    cheek_left_mean=crop_img_left_cheek.mean()
    d['cheek_left_mean']=cheek_left_mean
    if xb + w - xre < Le / 2:  #Right Cheek
        rect = cv2.boundingRect(pts_right_cheek)
        x, y, w, h = rect
        crop_img_right_cheek = thermal_image[y:y + h, x:x + w].copy()  # Need to make more changes.
    else:
        yrc3 = int(round(xre + Le / 2, 0))
        pts_right_cheek = np.array([[xre, ylc2], [xrm, yram], [yrc3, yram], [yrc3, ylc2]], np.int32)
        pts_right_cheek = pts_right_cheek.reshape((-1, 1, 2))
        rect = cv2.boundingRect(pts_right_cheek)
        x, y, w, h = rect
        crop_img_right_cheek = thermal_image[y:y + h, x:x + w].copy()  # Need to make more changes.
    cheek_right_mean = crop_img_right_cheek.mean()
    d['cheek_right_mean'] = cheek_right_mean
    #Forehead
    rect = cv2.boundingRect(pts_forhead)
    x, y, w, h = rect
    crop_img_forehead = thermal_image[y:y + h, x:x + w].copy()  # Need to make more changes.
    d['forehead']=crop_img_forehead.mean()
    #cv2.imshow("image", thermal_image)
    cv2.waitKey(0)
    return d
#To extract temperatures from the images captured using flir lepton
# import flirimageextractor
# from matplotlib import cm
#
# flir = flirimageextractor.FlirImageExtractor(palettes=[cm.jet, cm.bwr, cm.gist_ncar])
# flir.process_image('Test1/Thermal (10).jpg')
# temp=flir.get_thermal_np
# print(temp)
# DYLD_PRINT-_LIBRARIES=1
# To read each file on rgb, thermal to get the intensity values
i=0
for filename in glob.glob('Test1/RGB*.jpg'):
        print(filename)
        number=filename[filename.find("(")+1:filename.find(")")] #Gives the RGB image number which is in the file name
        #print(number)
        thermal_filename='Test1\Thermal ('+number+').jpg' #To read corresponding thermal image file
        #print(thermal_filename)
        thermal_image=cv2.imread(thermal_filename) #Reading thermal image
        thermal_image = cv2.resize(thermal_image, None, fx=4, fy=4) #Rescaling the thermal image to match RGB dimensions
        #Gaussian filter
        thermal_gaussian = cv2.GaussianBlur(thermal_image, (3, 3), 0)
        #median filter
        thermal_median = cv2.medianBlur(thermal_image, 3)
        image = cv2.imread(filename) #Reding RGB image
        #cv2.imshow('image', image) #To show the readed file not needed
        result = detector.detect_faces(image) #Running CNN to detect regions

        if result:
                intensity_mean=output(result,thermal_gaussian)
                #print(intensity_mean)
        # bounding_box = result[0]['box'] #To get the x,y,w,h of the face
        # keypoints = result[0]['keypoints']
        # crop_img_face = thermal_image[bounding_box[1]-30:bounding_box[1]-30 + bounding_box[3], bounding_box[0]:bounding_box[0]+bounding_box[2]] #Cropping thermal image based on the face dimensions
        #
        # print(crop_img_face.mean())  #To print the mean value of the cropped image
        #Storing data in excel
                sheet.cell(row=i+2, column=1).value = number
                sheet.cell(row=i+2, column=2).value = intensity_mean['face']
                sheet.cell(row=i+2, column=3).value = intensity_mean['cheek_left_mean']
                sheet.cell(row=i+2, column=4).value = intensity_mean['cheek_right_mean']
                sheet.cell(row=i+2, column=5).value = intensity_mean['forehead']
                sheet.cell(row=i+2, column=6).value = intensity_mean['nose_mean']
                i = i + 1
wb.save('intensity_values_eachROI_gaussian.xlsx') #To save the workbook

#image = cv2.imread("Test1/RGB (10).jpg")
#thermal_filename='Test1/Thermal (10) .jpg'
#thermal_img = cv2.imread('Test1/Thermal (10).jpg')
# cv2.imshow('image', image)
# print(image.mean())
# img = cv2.imread("Test1/RGB (100).jpg")
# img_1 = cv2.imread("Test1/Thermal (100).jpg")
#thermal_img = cv2.resize(thermal_img,None,fx=4,fy=4)
#result = detector.detect_faces(image)
# print(result)
#mean=output(result,thermal_img)
#print(mean)
# keypoints = result[0]['keypoints']
#bounding_box = result[0]['box']
#crop_img = thermal_img[bounding_box[1]-30:bounding_box[1]-30 + bounding_box[3], bounding_box[0]:bounding_box[0]+bounding_box[2]]
#cv2.rectangle(thermal_img,
                    # (bounding_box[0], bounding_box[1]-30),
                    # (bounding_box[0]+bounding_box[2], bounding_box[1]-30 + bounding_box[3]),
                    # (0,155,255),
                    # 2)
#cv2.rectangle(image,
                    # (bounding_box[0], bounding_box[1]),
                    # (bounding_box[0]+bounding_box[2], bounding_box[1] + bounding_box[3]),
                    # (0,155,255),
                    # 2)
#cv2.imshow("thermal", thermal_img)
#cv2.imshow("RGB",image)
#cv2.imwrite('Test1/result/RGB(10).png',image)
#cv2.imwrite('Test1/result/Thermal(10).png',thermal_img)
# print(crop_img.mean())
cv2.waitKey(0)
# image = cv2.imread("Test1/RGB (10).jpg")
# base=os.path.basename('Test1/RGB (10).jpg')
# print(base)
# base2=os.path.splitext(base)
# print(base2)
# number=base[base.find("(")+1:base.find(")")]
# print(number)
# thermal_image=cv2.imread("Test1/Thermal (10).jpg")
# thermal_image = cv2.resize(thermal_image,None,fx=4,fy=4)
# #cap = cv2.VideoCapture('for_testing.mov')
# import numpy as np
# result = detector.detect_faces(image)
# print(result)
# bounding_box = result[0]['box']
# keypoints = result[0]['keypoints']
# cv2.rectangle(image,
#                     (bounding_box[0], bounding_box[1]),
#                     (bounding_box[0]+bounding_box[2], bounding_box[1] + bounding_box[3]),
#                     (0,155,255),
#                     2)
# cv2.rectangle(thermal_image,
#                     (bounding_box[0], bounding_box[1]-30),
#                     (bounding_box[0]+bounding_box[2], bounding_box[1]-30 + bounding_box[3]),
#                     (0,155,255),
#                     2)
# cv2.circle(image,(keypoints['nose']), 2, (0,155,255), 2)
# cv2.circle(thermal_image,(keypoints['nose']), 2, (0,155,255), 2)
# cv2.namedWindow("image")
# cv2.imshow("image",image)
# cv2.imshow("thermal",thermal_image)
# cv2.waitKey(0)